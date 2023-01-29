import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import utils
from copy import copy
import argparse


# コマンドラインから引数の受け取り，ブックの作成と読み込み
psr  = argparse.ArgumentParser(description='第一引数にASPで導出したファイル，第二引数に設定用ファイルを指定してください')
psr.add_argument('arg1')
psr.add_argument('arg2')
args = psr.parse_args()

book = op.load_workbook(args.arg1)
setting_book = op.load_workbook(args.arg2)
#new_book = op.Workbook()

# 設定用ファイルの希望シフトシートを取得
sbs = setting_book.worksheets[3]
sbs.insert_cols(5)
# シートを取得しコピー，新たなシートを作成
ws = book.worksheets[0]

#ws1 = book.create_sheet(index=1, title='希望シフトの確認')
ws1 = book.copy_worksheet(book.worksheets[0]) # 希望シフト
ws2 = book.copy_worksheet(book.worksheets[0]) # N,J
ws3 = book.copy_worksheet(book.worksheets[0]) # 日勤
ws4 = book.copy_worksheet(book.worksheets[0]) # P
ws5 = book.copy_worksheet(book.worksheets[0]) # 夜勤
ws6 = book.copy_worksheet(book.worksheets[0]) # ○,◎
ws7 = book.copy_worksheet(book.worksheets[0]) # 休暇
#new_book = new_book.copy_worksheet(setting_book.worksheets[3])

# 結合されたセルをすべて解除
mg = []
for i in ws1.merged_cells.ranges :
    mg.append(i)
for i in mg:
    ws1.unmerge_cells(None,i.min_row,i.min_col,i.max_row,i.max_col)

# 名前の変更
ws1.title='希望シフト'
ws2.title = 'N,J'
ws3.title = '日勤'
ws4.title = 'P'
ws5.title = '夜勤'
ws6.title = '○,◎'
ws7.title = '休暇'

# すべてを色なしに
for row in ws1:
  for cell in row:
      cell.fill = PatternFill(fill_type = None)
      ws1[cell.coordinate].font = op.styles.fonts.Font(color='000000')

'''
# 希望シフトがある場合は色を塗る
for row in sbs['L3':'AM34']:
    #行データの確保
    sbs_cells = [v.value for v in row]
    for cell in row:
        if cell.value != None:
            ws1[cell.coordinate].fill = PatternFill(fgColor='ADFF2F',bgColor="ADFF2F", fill_type = "solid")
'''

# 間に空の行を挿入
for row_index in range(3,ws1.max_row):
    if row_index%2 == 1 and row_index != 1:
        insert_num = row_index + 1
        ws1.insert_rows(insert_num)
        

side1 = Side(style='thin', color='000000')
side2 = Side(style='thick', color='000000')
side3 = Side(style='dashed', color='000000')

border_thin = Border(top=side1,left=side1, right=side1)
border_thick = Border(top=side2)
border_dashed = Border(top=side3,left=side1, right=side1)
border_ws1 = Border(top=side2,left=side1, right=side1,bottom=None)
border_dashed_thick_left = Border(top=side3,left=side2, right=side1)
border_ws1_thick_left = Border(top=side2,left=side2, right=side1,bottom=None)
#border_column = Border(left=side1, right=side1)
'''
# 各シフトの下に希望シフトを追記
for i in range(3, sbs.max_row):
    for j in range(1, 48):
        ws1.cell(row=i*2-3, column=j).value = ws.cell(row=i, column=j).value
        ws1.cell(row=i*2-3, column=j).border = border_thin
        #ws1.cell(row=i*2-2, column=j).value = sbs.cell(row=i, column=j).value
        ws1.cell(row=i*2-3, column=j).alignment = Alignment(horizontal = 'center', vertical = 'center')
        #if ws1.cell(row=i*2-3, column=j).value != sbs.cell(row=i, column=j).value and sbs.cell(row=i, column=j).value != None:
         #   ws1.cell(row=i*2-2, column=j).fill = PatternFill(fgColor='FF1493',bgColor="FF1493", fill_type = "solid")
'''

# 各シフトの下に希望シフトを追記
for i in range(3, 35):
    ws1.row_dimensions[i*2-2].height = 13 #高さ13に揃える
    for j in range(6, 48):
        ws1.cell(row=i*2-2, column=j).value = sbs.cell(row=i, column=j).value
        cell_string = str(ws1.cell(row=i*2-2, column=j).value)
        ws1.cell(row=i*2-2, column=j).alignment = Alignment(horizontal = 'center', vertical = 'center')
        if cell_string == "S":
            ws1.cell(row=i*2-2, column=j).value = "Ｓ"
        if cell_string == "J":
            ws1.cell(row=i*2-2, column=j).value = "Ｊ"
        if cell_string == "N":
            ws1.cell(row=i*2-2, column=j).value = "Ｎ"
        if cell_string == "P":
            ws1.cell(row=i*2-2, column=j).value = "Ｐ"
        if cell_string == "/":
            ws1.cell(row=i*2-2, column=j).value = "／"
        if j == 6 or j == 13 or j == 41 or j == 48:
            ws1.cell(row=i*2-3, column=j).border = border_ws1_thick_left
            ws1.cell(row=i*2-2, column=j).border = border_dashed_thick_left
        else:
            ws1.cell(row=i*2-3, column=j).border = border_ws1
            ws1.cell(row=i*2-2, column=j).border = border_dashed
        #if ws1.cell(row=i*2-3, column=j).value != None:
            #ws1.cell(row=i*2-3, column=j).fill = PatternFill(fgColor='FF1493',bgColor="FF1493", fill_type = "solid")
        if ws1.cell(row=i*2-2, column=j).value != None and 12 < j and j < 41:
            ws1.cell(row=i*2-3, column=j).fill = PatternFill(fgColor='79c06e',bgColor="FF1493", fill_type = "solid")
            if ws1.cell(row=i*2-2, column=j).value != ws1.cell(row=i*2-3, column=j).value:
                ws1.cell(row=i*2-3, column=j).fill = PatternFill(fgColor='ea553a',bgColor="FF1493", fill_type = "solid")
                ws1.cell(row=i*2-2, column=j).fill = PatternFill(fgColor='ea553a',bgColor="FF1493", fill_type = "solid")
        #ws1.cell(row=i*2-2, column=j).value = op.styles.FORMAT_TEXT 
        #cell_value = ws1.cell(row=i*2-2, column=j).value
        #cell_value.translate(str.maketrans({chr(0x0021 + i): chr(0xFF01 + i) for i in range(94)}))
        #ws1.cell(row=i*2-2, column=j).value = cell_value
        #if ws1.cell(row=i*2-3, column=j).value != sbs.cell(row=i, column=j).value and sbs.cell(row=i, column=j).value != None:
            #ws1.cell(row=i*2-2, column=j).fill = PatternFill(fgColor='FF1493',bgColor="FF1493", fill_type = "solid")

#無駄な行の削除
Sheet_Max = ws1.max_row+1
stop_row = 67

for row in reversed(range(stop_row,Sheet_Max)):
    ws1.delete_rows(row)

'''
for row in sbs:
    for cell in row:
        ws1[cell.coordinate].value = cell.value

for row in ws:
    for cell in row:
        ws2[cell.coordinate].value = cell.value
'''

def No_Color(sheet):
    for row in sheet:
        for cell in row:
            if cell.row > 2:
                cell.fill = PatternFill(fill_type = None)
                sheet[cell.coordinate].font = op.styles.fonts.Font(color='000000')

No_Color(ws2)
No_Color(ws3)
No_Color(ws4)
No_Color(ws5)
No_Color(ws6)
No_Color(ws7)

def Coloring(sheet,shift,color):
    # NまたはJの場合は色を塗る
    for row in sheet['M3':'AN34']:
        for cell in row:
            if cell.value == shift:
                cell.fill = PatternFill(fgColor=color,bgColor=color, fill_type = "solid")

Coloring(ws2,'Ｎ','ffaf00')
Coloring(ws2,'Ｊ','87ffd7')

Coloring(ws3,'日','ffcc66')

Coloring(ws4,'Ｐ','f6b483')

Coloring(ws5,'Ｊ','87ffd7')
Coloring(ws5,'Ｓ','00afaf')
Coloring(ws5,'★','87afff')
Coloring(ws5,'☆','8787ff')

Coloring(ws6,'◎','a4a8d4')
Coloring(ws6,'○','a4a8d4')

Coloring(ws7,'◎','a4a8d4')
Coloring(ws7,'○','a4a8d4')
Coloring(ws7,'年','e3acae')
Coloring(ws7,'健','e3acae')

def Coloring_day_offs(sheet,shift,color):
    # NまたはJの場合は色を塗る
    for row in sheet['M3':'AN34']:
        for cell in row:
            if cell.value == shift :
                cell.fill = PatternFill(fgColor=color,bgColor=color, fill_type = "solid")


# 保存する
book.save('new-4n.xlsx')