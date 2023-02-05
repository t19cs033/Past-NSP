# -*- coding: utf-8 -*-
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.numbers import builtin_format_code, builtin_format_id
from openpyxl.styles.numbers import is_builtin, is_date_format, is_datetime 
from openpyxl import utils
from copy import copy
import argparse
import datetime

dist_sheet_num = 40
# コマンドラインから引数の受け取り，ブックの作成と読み込み
psr  = argparse.ArgumentParser(description='')
psr.add_argument('arg1')
psr.add_argument('arg2')
args = psr.parse_args()

past_data_book = op.load_workbook(args.arg1)
setting_book = op.load_workbook(args.arg2)
#new_book = op.Workbook()
pdb_1 = past_data_book.worksheets[dist_sheet_num-1] #2022-05-01(35)
pdb_2 = past_data_book.worksheets[dist_sheet_num] #2022-05-29(36)
pdb_3 = past_data_book.worksheets[dist_sheet_num+1] #2022-06-26(37)
sb = setting_book.worksheets[3] #希望シフトがあるシート

#置換する文字列を指定
Henkan_mae = ['期間：',' ～ ４週間','年','月','日']
Henkan_go = ['','','/','/','']
i = 0

for list in Henkan_mae:
    i = i + 1
    #置換
    new_text = pdb_2.cell(row=1,column=1).value.replace(list, Henkan_go[i-1])
    pdb_2.cell(row=1,column=1).value = new_text

datetime_text = new_text.split('/')
date = datetime.date(int(datetime_text[0]), int(datetime_text[1]), int(datetime_text[2]))
sb.cell(row=1,column=2).value = date

date_int = date - datetime.date(2022,1,1)
#print(date_int.days)

str_text = str(new_text)
replace_text = str_text.replace('/','-')
facts_name = replace_text + ".lp"

#新規テキストファイル作成
f = open(facts_name, 'w')

# 希望シフトを吐く関数
def PrintShifts(sheet, startcol, endcol, diffcol):
    for i in range(5, 89):
        if i%2 == 1: #元の勤務表が2行1名の記述なので1行飛ばしでループ
            for j in range(startcol, endcol):
                cell_string = str(sheet.cell(row=i, column=j).value) #cell_stringはセルに存在する文字列
                if cell_string == "○":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "○"
                if cell_string == "日◇":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "日"
                if cell_string == "J2S":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "J"
                if cell_string == "SS":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "S"
                if cell_string == "N2K":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "N"
                if cell_string == "N2":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "N"
                if cell_string == "J2":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "J"
                if cell_string == "Ｓ":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "S"
                if cell_string == "年":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "年"
                if cell_string == "健":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "健"
                if cell_string == "★":  
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "★"
                if cell_string == "☆":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "☆"
                if cell_string == "P4":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "P"
                if cell_string == "◎":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "◎"
                if cell_string == "" or cell_string == "－" or cell_string == "パ7":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "／"
                if cell_string == "□":
                    sb.cell(row=(i+1)/2, column=j+diffcol).value = "特"
# shift:転記元のシート名
# startcol:転記元の何列目から読み取るか
# diffcol:転記先の何列目から記載するか

# assignedを吐く関数
def PrintAssigned(sheet, startcol, endcol, diffday):
    staff_id = 0

    for i in range(5, 89):
        if i%2 == 1: #元の勤務表が2行1名の記述なので1行飛ばしでループ
            staff_id += 1 #行ごとにスタッフのIDを更新
            day_id = date_int.days + diffday #date_int.daysは開始日の日付の整数値,1/1を引いているので+１している
            for j in range(startcol, endcol):
                day_id += 1 #列が移動するごとに日付も更新する
                cell_string = str(sheet.cell(row=i, column=j).value) #cell_stringはセルに存在する文字列
                if cell_string == "○":
                    f.write("assigned(" + str(staff_id) + "," + str(day_id) + ",\"○\").\n") #assigned(N,D,S)
                if cell_string == "日◇":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"日\").\n") #assigned(N,D,S)
                if cell_string == "J2S":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"J\").\n") #assigned(N,D,S)
                if cell_string == "SS":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"S\").\n") #assigned(N,D,S)
                if cell_string == "N2K":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"N\").\n") #assigned(N,D,S)
                if cell_string == "N2":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"N\").\n") #assigned(N,D,S)
                if cell_string == "J2":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"J\").\n") #assigned(N,D,S)
                if cell_string == "Ｓ":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"S\").\n") #assigned(N,D,S)
                if cell_string == "年":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"年\").\n") #assigned(N,D,S)
                if cell_string == "健":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"健\").\n") #assigned(N,D,S)
                if cell_string == "★":  
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"★\").\n") #assigned(N,D,S)
                if cell_string == "☆":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"☆\").\n") #assigned(N,D,S)
                if cell_string == "P4":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"P\").\n") #assigned(N,D,S)
                if cell_string == "◎":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"◎\").\n") #assigned(N,D,S)
                if cell_string == "" or cell_string == "－" or cell_string == "パ7":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"/\").\n") #assigned(N,D,S)
                if cell_string == "□":
                    f.write("assigned("+str(staff_id)+","+str(day_id)+",\"特\").\n") #assigned(N,D,S)

PrintShifts(pdb_1,27,34,-22) #27列目を5列目に記載したいため，diffcol = 5-27 = -22
PrintShifts(pdb_2,6,34,6) #6列目を11列目に記載したいため，diffcol = 11-6+1 = 6
#PrintShifts(pdb_3,6,13,34) #6列目を40列目に記載したいため，diffcol = 40-6 = 34
PrintAssigned(pdb_1,27,34,-7) #27列目を5列目に記載したいため，diffcol = 5-27 = -22
PrintAssigned(pdb_2,6,34,0) #6列目を11列目に記載したいため，diffcol = 11-6+1 = 6
#PrintAssigned(pdb_3,6,13,34) #6列目を40列目に記載したいため，diffcol = 40-6 = 34

for row in sb['B35':'AT50']:
    for cell in row:
        cell.value = None
        cell.number_format = builtin_format_code(0) #セルのフォーマットを標準へ変更

file_name = "4n-" + replace_text + ".xlsx"
# 保存する
setting_book.save(file_name)
f.close()